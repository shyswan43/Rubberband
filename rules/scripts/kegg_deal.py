import os
import re
from copy import deepcopy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.rich_text import TextBlock
from openpyxl.utils import get_column_letter


def count_keywords_in_txt(txt_path, keyword):
    """
    统计txt文件中关键词出现的次数
    """
    try:
        with open(txt_path, 'r', encoding='utf-8') as f:
            text = f.read()
            # 使用正则表达式匹配关键词（支持中文、英文、数字）
            pattern = re.compile(re.escape(keyword))
            return len(pattern.findall(text))
    except Exception as e:
        print(f"读取文件 {txt_path} 时出错: {e}")
        return 0

def merge_xlsx_with_txts():
    """
    复制原xlsx文件并添加统计结果，保留所有格式
    """
    # 1. 定义文件夹路径（根据实际情况修改）
    txt_folder = snakemake.params.input_dir  # 包含txt文件的文件夹路径
    output_path = os.path.join(txt_folder,"summary_kegg_sp.xlsx")  # 输出文件名
    output_fn = os.path.join(txt_folder,"summary_kegg.xlsx")

    # 2. 复制原文件（保留格式）
    #new_wb = load_workbook(output_path,keep_vba=True,data_only=False,rich_text=True)  # 直接加载原文件作为副本

    ##富文本处理
 # 1. 加载源工作簿（必须设置data_only=False以保留公式和格式）
    wb_source = load_workbook(output_path, data_only=False, rich_text=True)
    new_wb = load_workbook(output_path, data_only=False, rich_text=True)  # 基于源文件创建目标工作簿
    
    # 2. 遍历源工作簿的所有工作表
    for sheet_name in wb_source.sheetnames:
        ws_source = wb_source[sheet_name]
        ws_target = new_wb.create_sheet(title=sheet_name)  # 在目标工作簿中创建同名工作表
        
        # 3. 复制每个单元格的内容和富文本格式
        for row in ws_source.iter_rows():
            for cell in row:
                # 复制单元格值（自动处理富文本）
                ws_target[cell.coordinate].value = cell.value
                
                # 复制单元格样式（字体、填充、边框、对齐等）
                if cell.has_style:
                    ws_target[cell.coordinate]._style = deepcopy(cell._style)
                
                # 特殊处理富文本：确保Run对象的字体属性被正确复制
                if hasattr(cell, 'rich_text') and cell.rich_text:
                    target_rich_text = []
                    for run in cell.rich_text:
                        new_run = deepcopy(run)
                        # 显式复制字体属性（避免引用导致的样式丢失）
                        new_run.font = deepcopy(run.font)
                        target_rich_text.append(new_run)
                    ws_target[cell.coordinate].rich_text = target_rich_text
        
        # 4. 复制行高和列宽（保持表格布局一致）
        for col in ws_source.column_dimensions:
            ws_target.column_dimensions[col].width = ws_source.column_dimensions[col].width
        for row in ws_source.row_dimensions:
            ws_target.row_dimensions[row].height = ws_source.row_dimensions[row].height
        
        # 5. 复制合并单元格（保持表格结构一致）
        for merged_range in ws_source.merged_cells.ranges:
            ws_target.merge_cells(
                start_row=merged_range.min_row,
                start_column=merged_range.min_col,
                end_row=merged_range.max_row,
                end_column=merged_range.max_col
            )
    ##富文本处理

    # 3. 获取txt文件列表（不含扩展名）
    txt_files = [f[:-4] for f in os.listdir(txt_folder) if f.endswith('.txt')]
    if not txt_files:
        print("目标文件夹中没有找到txt文件")
        return

    # 4. 在第八列及之后添加数据
    for sheet in new_wb.sheetnames:
        ws = new_wb[sheet]
        txt_col = 8  # 第八列的列号（openpyxl中列索引从1开始）
        
        # 在第一行写入txt文件名（从第八列开始）
        ws.cell(row=1, column=txt_col, value=txt_files[0])  # 第八列第一行
        for i, txt_file in enumerate(txt_files[1:], start=1):
            ws.cell(row=1, column=txt_col + i, value=txt_file)  # 后续列第一行
        
        # 从第二行开始处理
        for row in range(2, ws.max_row + 1):
            keyword = ws.cell(row=row, column=5).value  # 第五列的数据作为关键词
            if not keyword or keyword == '':
                continue
                
            # 从第八列开始处理每一列
            for i, txt_file in enumerate(txt_files, start=0):
                col = txt_col + i
                # 构建txt文件路径
                full_txt_path = os.path.join(txt_folder, txt_file + '.txt')
                if os.path.exists(full_txt_path):
                    # 统计关键词出现次数
                    count = count_keywords_in_txt(full_txt_path, keyword)
                else:
                    count = 0
                ws.cell(row=row, column=col, value=count)

    # 5. 保存结果
    new_wb.save(output_path)
    os.rename(output_path,output_fn)
    print(f"结果已保存到 {output_fn}")

if __name__ == "__main__":
    merge_xlsx_with_txts()